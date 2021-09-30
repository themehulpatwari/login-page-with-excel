
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment

workbook = openpyxl.load_workbook('login_python.xlsx') # Importing excel file for changes
workbook2 = pd.read_excel('login_python.xlsx') # Just Reading the excel sheet
n = len(workbook2)
username = []
ws = workbook.worksheets[0] # Sheet number

for i in range(n):
    username.append(workbook2['USERNAME'].loc[i]) # List of username already in use

while True:
    new_username = input('Give a new username: ')
    if new_username in username:
        print('Username Taken, try different')
    else:
        break
new_password = input('Give a new password: ')
print('\nUsername and Password created successfully')

nc_username = ws.cell(row = n+2, column = 1)
nc_username.value = new_username # Writing the username in the blank cell

nc_password = ws.cell(row = n+2, column = 2)
nc_password.value = new_password # Writing the password in the blank cell

# Aligning the cell, because it looks good and I like it.

nc_username.alignment = Alignment(horizontal = 'center')
nc_password.alignment = Alignment(horizontal = 'center')

workbook.save('login_python.xlsx') # Saving the workbook

user_login = input('Reply with either yes or no\nDo you want to login? ')

# Asking the user whether they want to login
if user_login == 'yes':
    print()
    import login_excel # Importing code for login

