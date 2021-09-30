import random
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
import os
import smtplib
from email.message import EmailMessage

def mail_otp(email,otp):
    EMAIL_ADDRESS = os.environ.get('EMAIL_USER') # Environment Variable containing email and password
    EMAIL_PASSWORD = os.environ.get('EMAIL_PASS')

    # Body of the email
    msg = EmailMessage()
    msg['Subject'] = 'One Time Password for email verification'
    msg['From'] = EMAIL_ADDRESS
    msg['To'] = email
    msg.set_content(('''Your OTP for verification is
        {}.''').format(otp))

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(EMAIL_ADDRESS,EMAIL_PASSWORD)

        smtp.send_message(msg)


workbook = openpyxl.load_workbook('login_python.xlsx') # Importing excel file for changes
workbook2 = pd.read_excel('login_python.xlsx') # Just Reading the excel sheet
n = len(workbook2)
username = []
ws = workbook.worksheets[0] # Sheet number

for i in range(n):
    username.append(workbook2['USERNAME'].loc[i]) # List of username already in use

while True:
    new_username = input('Give your email id for Sign Up: ')
    if new_username in username:
        print('Username Taken, try different')
    else:
        break

while True:
    otp = random.randint(1001,9999)
    mail_otp(new_username,otp) # Sending the otp though mail
    print('An otp has been sent to your email: {}, Please Verify the otp'.format(new_username))
    print()
    user_otp = input('Please type your otp here: ')
    if otp == int(user_otp):
        print('\nYour email has been verified. Thank You')
        break
    else:
        print('Wrong OTP, account craetion suspended')
        sys.exit()

new_password = input('Give a new password: ')
print('\nUsername and Password created successfully')

nc_username = ws.cell(row = n+2, column = 1)
nc_username.value = new_username # Writing the username in the blank cell

nc_password = ws.cell(row = n+2, column = 2)
nc_password.value = new_password # Writing the password in the blank cell

# Aligning the cell, because it looks good and I like it.

nc_username.alignment = Alignment(horizontal = 'center')
nc_password.alignment = Alignment(horizontal = 'center')

workbook.save('login_python.xlsx') # Saving the workbook
